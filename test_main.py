from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook

# Ensure project root is importable when running tests from repo root
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from main import ExtractConfig, process_workbook


def _make_en_general_sheet(wb: Workbook, include_key: str | None) -> None:
    ws = wb.create_sheet("TGS_EN_GNRL_ICG")
    # Minimal shape resembling the screenshot: a label cell and many values across the row.
    ws["B2"] = "Incl_Country_LOB_Lst"
    if include_key is not None:
        ws["L2"] = include_key


def _build_workbook(path: Path, with_country_tab: bool, include_key: str | None, with_change_log: bool = True) -> None:
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    default = wb.active
    wb.remove(default)

    if with_change_log:
        wb.create_sheet("Change Log")

    if with_country_tab:
        wb.create_sheet("TGS_EC_ICG")

    _make_en_general_sheet(wb, include_key)

    # Add a few other country sheets to ensure they're removed
    wb.create_sheet("TGS_AR_ICG")
    wb.create_sheet("TGS_BR_ICG")

    wb.save(str(path))


def test_keeps_change_log_country_and_en_general(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    output_dir.mkdir()

    src = input_dir / "Scenario_116000079.xlsx"
    _build_workbook(src, with_country_tab=True, include_key="LATAM_EC_ICG")

    cfg = ExtractConfig(input_dir=input_dir, output_dir=output_dir, country_code="EC", region="LATAM", lob="ICG")
    out_path = process_workbook(src, cfg)
    assert out_path is not None and out_path.exists()

    wb_out = load_workbook(str(out_path))
    names = set(wb_out.sheetnames)
    assert names == {"Change Log", "TGS_EC_ICG", "TGS_EN_GNRL_ICG"}


def test_keeps_only_en_general_when_no_country_tab_but_membership(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    output_dir.mkdir()

    src = input_dir / "Scenario_External.xlsx"
    _build_workbook(src, with_country_tab=False, include_key="LATAM_EC_ICG")

    cfg = ExtractConfig(input_dir=input_dir, output_dir=output_dir, country_code="EC", region="LATAM", lob="ICG")
    out_path = process_workbook(src, cfg)
    assert out_path is not None and out_path.exists()

    wb_out = load_workbook(str(out_path))
    names = set(wb_out.sheetnames)
    assert names == {"Change Log", "TGS_EN_GNRL_ICG"}


def test_writes_nothing_when_not_member_and_no_country_tab(tmp_path: Path) -> None:
    input_dir = tmp_path / "in"
    output_dir = tmp_path / "out"
    input_dir.mkdir()
    output_dir.mkdir()

    src = input_dir / "Scenario_Not_Relevant.xlsx"
    # No country tab; membership list does not include LATAM_EC_ICG
    _build_workbook(src, with_country_tab=False, include_key="LATAM_AR_ICG")

    cfg = ExtractConfig(input_dir=input_dir, output_dir=output_dir, country_code="EC", region="LATAM", lob="ICG")
    out_path = process_workbook(src, cfg)
    assert out_path is None


