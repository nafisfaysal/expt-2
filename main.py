#!/usr/bin/env python3
"""
TGS Sheet Automation

This script processes TGS scenario Excel workbooks and extracts only the
tabs relevant to a target country (e.g., Ecuador -> "EC").

Rules distilled from the provided document:
  - Workbooks are built per scenario, not per country.
  - Keep the "Change Log" tab in the output if it exists.
  - If a country-specific tab exists (e.g., "TGS_EC_ICG"), keep it
    and drop other country tabs.
  - Some scenarios (e.g., External Entity) use a shared general tab
    (e.g., "TGS_EN_GNRL_ICG"). Keep that general tab only if the
    row with TSHLD_NM == "Incl_Country_LOB_Lst" contains the country
    in its list (e.g., value includes "LATAM_EC_ICG").

Outputs a copy of each matching workbook into an output directory,
retaining only the relevant sheets.

Usage example:
  python main.py \
    --input-dir /path/to/scenarios \
    --output-dir /path/to/output \
    --country EC \
    --region LATAM \
    --lob ICG

Notes:
  - Supports .xlsx and .xlsm (macros are preserved if present).
  - Requires: openpyxl
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Set, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ----------------------------
# Configuration data structure
# ----------------------------


@dataclass(frozen=True)
class ExtractConfig:
    input_dir: Path
    output_dir: Path
    country_code: str  # e.g., "EC"
    region: str = "LATAM"  # e.g., "LATAM"
    lob: str = "ICG"  # e.g., "ICG"
    dry_run: bool = False

    def normalized_country(self) -> str:
        return self.country_code.strip().upper()

    def country_tab_name(self) -> str:
        return f"TGS_{self.normalized_country()}_{self.lob.upper()}"

    def en_general_tab_name(self) -> str:
        return f"TGS_EN_GNRL_{self.lob.upper()}"

    def region_country_key(self) -> str:
        """Returns the membership key used in Incl_Country_LOB_Lst."""
        return f"{self.region.upper()}_{self.normalized_country()}_{self.lob.upper()}"


# ----------------------------
# Helpers
# ----------------------------


CHANGE_LOG_CANDIDATES: Tuple[str, ...] = (
    "Change Log",
    "ChangeLog",
    "CHANGE LOG",
    "CHANGE_LOG",
)


def normalize_string(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    return (
        text.strip()
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\xa0", " ")
    )


def sheet_exists(sheet_names: Sequence[str], name: str) -> bool:
    name_norm = name.strip().lower()
    return any(s.strip().lower() == name_norm for s in sheet_names)


def find_change_log_sheet(sheet_names: Sequence[str]) -> Optional[str]:
    for candidate in CHANGE_LOG_CANDIDATES:
        if sheet_exists(sheet_names, candidate):
            return next(s for s in sheet_names if s.strip().lower() == candidate.strip().lower())
    return None


def contains_country_in_en_general(ws: Worksheet, membership_key: str) -> bool:
    """Return True if the EN general sheet includes the country in Incl_Country_LOB_Lst.

    Strategy:
      1) Try to locate a cell whose normalized value == "incl_country_lob_lst".
         If found, search that row for the membership_key.
      2) Fallback: scan entire sheet for membership_key substring.
    """

    target_label = "incl_country_lob_lst"
    membership_key_norm = membership_key.strip().upper()

    # 1) Targeted scan by label
    try:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                value = normalize_string(cell.value)
                if value and value.replace(" ", "").replace("-", "_").replace("__", "_").lower() == target_label:
                    # Consider entire row's textual content for membership
                    row_text = " | ".join(normalize_string(c.value) for c in row)
                    if membership_key_norm in row_text.upper():
                        return True
                    # Also look to the right in a wider window (some files put the list several columns away)
                    max_col = ws.max_column
                    values_right: List[str] = []
                    for c in range(cell.column, max_col + 1):
                        values_right.append(normalize_string(ws.cell(row=cell.row, column=c).value))
                    if membership_key_norm in " | ".join(values_right).upper():
                        return True
                    # Not found in this row; keep scanning other rows
    except Exception:
        # Be permissive; fall back to broad scan
        pass

    # 2) Broad scan across all cells (slower but robust)
    try:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value and membership_key_norm in str(value).upper():
                    return True
    except Exception:
        pass

    return False


def compute_sheets_to_keep(sheet_names: Sequence[str], ws_en_general: Optional[Worksheet], cfg: ExtractConfig) -> Set[str]:
    keep: Set[str] = set()

    # Always try to keep change log if present
    change_log = find_change_log_sheet(sheet_names)
    if change_log:
        keep.add(change_log)

    # Country-specific tab
    country_tab = cfg.country_tab_name()
    if sheet_exists(sheet_names, country_tab):
        # Use the canonical casing from the workbook
        keep.add(next(s for s in sheet_names if s.strip().lower() == country_tab.strip().lower()))

    # EN general tab conditionally
    en_general_tab = cfg.en_general_tab_name()
    if sheet_exists(sheet_names, en_general_tab) and ws_en_general is not None:
        if contains_country_in_en_general(ws_en_general, cfg.region_country_key()):
            keep.add(next(s for s in sheet_names if s.strip().lower() == en_general_tab.strip().lower()))

    return keep


def process_workbook(path: Path, cfg: ExtractConfig) -> Optional[Path]:
    """Process a single Excel workbook and save a filtered copy if relevant.

    Returns the path to the output file if a filtered workbook was written,
    otherwise returns None (e.g., no relevant sheets for the target country).
    """

    if not path.suffix.lower() in {".xlsx", ".xlsm"}:
        return None

    # keep_vba=True conserves macros if the file is .xlsm
    keep_vba = path.suffix.lower() == ".xlsm"
    try:
        wb = load_workbook(filename=str(path), read_only=False, data_only=False, keep_vba=keep_vba)
    except Exception as exc:
        print(f"[WARN] Skipping {path.name}: failed to open workbook ({exc})")
        return None

    sheet_names: List[str] = list(wb.sheetnames)
    ws_en: Optional[Worksheet] = None
    if sheet_exists(sheet_names, cfg.en_general_tab_name()):
        ws_en = wb[cfg.en_general_tab_name()]

    keep = compute_sheets_to_keep(sheet_names, ws_en, cfg)

    # If we do not keep any country-relevant tab (neither country-specific nor EN general),
    # do not output.
    country_relevant = any(
        s.lower() in {cfg.country_tab_name().lower(), cfg.en_general_tab_name().lower()} for s in keep
    )
    if not country_relevant:
        return None

    if cfg.dry_run:
        print(f"[DRY-RUN] {path.name}: would keep -> {sorted(keep)}")
        return path  # indicate relevance without writing

    # Remove all other sheets
    to_remove = [s for s in sheet_names if s not in keep]
    for sheet in to_remove:
        try:
            wb.remove(wb[sheet])
        except Exception as exc:
            print(f"[WARN] {path.name}: failed to remove sheet '{sheet}': {exc}")

    # Ensure output dir exists
    cfg.output_dir.mkdir(parents=True, exist_ok=True)

    # Compose output file name
    out_name = f"{path.stem}__{cfg.normalized_country()}{path.suffix}"
    out_path = cfg.output_dir / out_name

    try:
        wb.save(str(out_path))
        print(f"[OK] Wrote {out_path.name}: kept {sorted(keep)}")
        return out_path
    except Exception as exc:
        print(f"[ERROR] Failed to save '{out_path.name}': {exc}")
        return None


def iter_workbooks(input_dir: Path) -> Iterable[Path]:
    for p in sorted(input_dir.glob("**/*")):
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"}:
            yield p


def parse_args(argv: Optional[Sequence[str]] = None) -> ExtractConfig:
    parser = argparse.ArgumentParser(description="Extract country-relevant sheets from TGS workbooks")
    parser.add_argument("--input-dir", required=True, type=Path, help="Directory containing scenario workbooks (.xlsx/.xlsm)")
    parser.add_argument("--output-dir", required=True, type=Path, help="Directory to write filtered copies")
    parser.add_argument("--country", required=True, help="2-letter country code, e.g., EC for Ecuador")
    parser.add_argument("--region", default="LATAM", help="Region prefix used in Incl_Country_LOB_Lst (default: LATAM)")
    parser.add_argument("--lob", default="ICG", help="LOB suffix used in sheet names (default: ICG)")
    parser.add_argument("--dry-run", action="store_true", help="Do not write files; only print what would be kept")

    args = parser.parse_args(argv)
    return ExtractConfig(
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        country_code=args.country,
        region=args.region,
        lob=args.lob,
        dry_run=args.dry_run,
    )


def main(argv: Optional[Sequence[str]] = None) -> int:
    cfg = parse_args(argv)

    if not cfg.input_dir.exists():
        print(f"[ERROR] Input dir does not exist: {cfg.input_dir}")
        return 2

    matched = 0
    written = 0
    for path in iter_workbooks(cfg.input_dir):
        result = process_workbook(path, cfg)
        if result is not None:
            matched += 1
            # In dry-run, result is the input path; in write-mode, it's the output path
            if not cfg.dry_run:
                written += 1

    if matched == 0:
        print("[INFO] No relevant scenarios found for the specified country.")
    else:
        if cfg.dry_run:
            print(f"[INFO] {matched} relevant workbook(s) would be produced (dry-run)")
        else:
            print(f"[INFO] Wrote {written} filtered workbook(s) for country {cfg.normalized_country()}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())


